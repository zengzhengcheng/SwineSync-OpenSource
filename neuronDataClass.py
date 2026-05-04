import sys
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

from datetime import datetime
import threading
import random
import gc
import concurrent.futures
from scipy import signal
import numpy as np
import pandas as pd
import pywt
import os
import h5py
import time
# 获取系统当前时区（例如 Asia/Shanghai）
from tzlocal import get_localzone
import pytz
from utils import getjilabel
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from pathlib import Path
import matplotlib.pyplot as plt
from models.ECGFeatureExtractor import ECGFeatureExtractor
from heartCorrect.rr_correction_continous import RRIntervalCorrector

class NeuronDataClass:
    def __init__(self,filepath,savepath,out=1,cache=False,cleanfile=False,correct=False,origin_resample=True):
        self.filepath=filepath
        self.savepath=savepath
        self.index = 0
        self.samplerate=512
        self.neuronLength=5120*3
        self.out=out
        self.meta=None
        self.cache=cache
        self.ready_event = threading.Event()  # 创建事件对象
        self.ready = False
        self.read_error = None
        self.plot_done=False
        self.cleanfile=cleanfile
        self.ye=1
        self.correct=correct
        self.corrector = RRIntervalCorrector(sampling_rate=512) if correct else None
        # origin_resample: True 表示保留原始时间戳（原位插值），False 表示均匀网格重采样
        self.origin_resample=origin_resample

    def checkzhen(self,threshold=245760):
        self.readData()
        df = pd.DataFrame({'timestamp': self.timestamp})
        df['timestamp'] = pd.to_datetime(df['timestamp'])
        df.set_index('timestamp', inplace=True)
        # 按10分钟重采样计数
        resampled = df.resample('10T').size().reset_index(name='count')
        resampled = resampled.iloc[1:-1]
        # 标记低于阈值的数据段
        resampled['warning'] = np.where(resampled['count'] < threshold, '数据不足', '')
        # 生成Excel文件并设置格式
        wb = Workbook()
        ws = wb.active

        # 写入数据
        for r in dataframe_to_rows(resampled, index=False, header=True):
            ws.append(r)

        # 设置标题样式
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill

        # 设置警告行的红色背景
        warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for row in range(2, len(resampled) + 2):
            if ws.cell(row=row, column=3).value == '数据不足':
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = warning_fill

        # 输出警告信息
        warning_periods = resampled[resampled['count'] < threshold]
        if not warning_periods.empty:
            print("\n警告：以下时间段数据量不足:")
            for _, row in warning_periods.iterrows():
                print(f"{row['timestamp']} - 数据量: {row['count']} (低于阈值 {threshold})")
        wb.save(self.savepath)

        return resampled
    def convertTocache(self):
        self.filename = os.path.basename(self.filepath)
        filename=Path(self.filename).stem+ "cache.csv"
        self.readRawTxt()
        if (self.cache):
            savefile = os.path.join(self.savepath, filename)
            print(f"  {datetime.now()}:正在生成缓存文件: {savefile} ...")
            # 创建DataFrame，第一列时间，第二列数据

            # 优化：直接创建DataFrame，减少中间步骤
            df = pd.DataFrame({
                'timestamp': self.timestamp,
                'data': self.data
            })
            
            # 清理实例变量
            self.timestamp = None
            self.data = None
            self.cleandata = None
            self.labels = None
            
            # 强制垃圾回收
            gc.collect()
            
            # 保存为CSV文件
            chunksize = 100000
            df.to_csv(savefile, index=False, sep=',', chunksize=chunksize)
            print(f"{savefile}缓存保存完成: {datetime.now()}")
            
            # 清理DataFrame
            del df
            gc.collect()
        
        # 彻底清理所有实例变量
        self.timestamp = None
        self.data = None
        self.cleandata = None
        self.labels = None
        self.filename = None
        self.filepath = None
        self.savepath = None
        self.length = 0
        self.index = 0
        self.indexlen = 0
        
        # 强制垃圾回收
        gc.collect()
        
        print(f"{datetime.now()}:缓存生成完成，内存已清理")
    def readData(self):
        self.read_error = None
        try:
            self.filename = os.path.basename(self.filepath)
            if(self.cache):
                self.load_from_cache(self.filepath)
            else:
                self.readRawTxt()
            self.ready = True
        except Exception as e:
            self.read_error = e
            self.ready = False
            print(f"[ERROR] readData 出错 ({self.filepath}): {e}")
        finally:
            self.ready_event.set()  # 无论成功或异常，始终触发事件，防止等待方永久阻塞
    def load_from_cache(self, filepath):
        # 如果缓存文件不存在，返回 False，代表加载失败
        if not os.path.exists(filepath):
            return False
        print(f"检测到缓存文件，正在快速读取: {filepath} {datetime.now()}")
        try:
            df = pd.read_csv(filepath, sep=',')
            # 兼容旧缓存（列名可能是 'ADC' 或第二列）
            if 'data' in df.columns:
                self.data = df['data'].values
            elif 'ADC' in df.columns:
                self.data = df['ADC'].values
            else:
                self.data = df.iloc[:, 1].values
            self.cleandata=np.zeros((len(self.data)))
            self.timestamp = df['timestamp'].values

            # 3. 恢复类的其他属性 (参照 readRawTxt 结尾的逻辑)
            self.length = len(self.data)
            self.labels = [0] * self.length
            self.calculateIndexlen()
            print(f"缓存读取完成 {self.filename} {datetime.now()}")
            self.ready_event.set()
            self.ready = True
            return True

        except Exception as e:
            print(f"缓存读取出错 (将重新读取原文件): {e}")
            return False
    def calculateIndexlen(self):
        if (self.ye == 1):
            indexlen = len(self.data) / 5120
            if (self.length % 5120 > 0):
                self.indexlen = int(indexlen) + 1
            else:
                self.indexlen = int(indexlen)
        else:
            indexlen = (len(self.data)) / self.neuronLength
            if (self.length % self.neuronLength > 0):
                self.indexlen = int(indexlen) + 1
            else:
                self.indexlen = int(indexlen)
    def readRawTxt(self):
        if("Raw" in self.filename or ("BMD101" in self.filename)):
            print(rf"read {self.filename} {datetime.now()}")
            # 不指定数据类型，避免NA值导致的错误
            if("Raw" in self.filename):
                rawdata = pd.read_csv(self.filepath, sep='\s+', header=None, names=["timestamp", "ADC"],
                                      skip_blank_lines=True, dtype_backend='pyarrow')
            else:
                rawdata = pd.read_csv(self.filepath, skiprows=1,sep=',', header=None, names=["timestamp", "ADC"],
                                      usecols=[0, 1], skip_blank_lines=True, dtype_backend='pyarrow')
            
            # 优化：及时清理数据
            rawdata.dropna(inplace=True)
            rawdata = rawdata.reset_index(drop=True)

            date=self.filename.split(".")[0]
            if("BMD101" in self.filename):
                date = date.split("_")[2]
            else:
                date = date.split("_")[1]
            year, month, day = map(int, date.split("-")[:3])
            starttime = datetime(year, month, day)
            
            # 向量化操作（比apply快10倍）
            parts = rawdata['timestamp'].str.replace('.', ':', regex=False).str.split(':', expand=True)
            parts = parts.reindex(columns=[0, 1, 2, 3])
            
            # 转换为数值类型（处理错误时返回 NaN）
            hours = pd.to_numeric(parts[0], errors='coerce',downcast='integer')
            minutes = pd.to_numeric(parts[1], errors='coerce',downcast='integer')
            seconds = pd.to_numeric(parts[2], errors='coerce',downcast='integer')
            millis = pd.to_numeric(parts[3], errors='coerce',downcast='integer')
            del parts
            gc.collect()
            
            # 计算总秒数（向量化操作，无循环）
            total_seconds = (
                    hours * 3600 +
                    minutes * 60 +
                    seconds +
                    millis.fillna(0) / 1000  # 处理毫秒缺失的情况
            )
            
            # 清理临时变量
            del hours, minutes, seconds, millis
            gc.collect()

            rawdata['timedelta'] = pd.to_timedelta(total_seconds, unit='s')
            
            # 清理临时变量
            del total_seconds
            gc.collect()
            
            rawdata.dropna(inplace=True)
            rawdata = rawdata.reset_index(drop=True)
            
            # 2. 自动检测跨天
            time_diff = rawdata['timedelta'].diff()
            cross_day = (time_diff < pd.Timedelta(0)).cumsum()
            
            # 清理临时变量
            del time_diff
            gc.collect()
            
            # 3. 生成最终时间戳（处理跨天）
            rawdata['timestamp'] = starttime + rawdata['timedelta'] + pd.to_timedelta(cross_day, unit='D')
            
            # 清理临时变量
            del cross_day
            gc.collect()
            
            rawdata['timestamp'] = pd.to_datetime(rawdata['timestamp'])
            rawdata['timestamp'] = rawdata['timestamp'].dt.strftime('%Y-%m-%d %H:%M:%S.%f').str[:-3]
            print(rf"完成时间校验 {self.filename} {datetime.now()}")
        else:
            # 不指定数据类型，避免NA值导致的错误
            rawdata = pd.read_csv(self.filepath, skiprows=1, sep='\s+',
                                  converters={'timestamp': lambda x: int(float(x.split(":")[0]))},
                                  names=["timestamp", "ADC", "HeartRate4sAverage", "HeartRate30sAverage"])
        
        # 优化：直接获取values，减少中间步骤
        self.timestamp = rawdata['timestamp'].values
        # 优化：使用向量化操作，减少内存使用
        self.data = (rawdata['ADC'].values.astype(np.float32) / 32767.0) * 100.0
        
        # 清理临时变量
        del rawdata
        gc.collect()
        
        self.cleandata = np.zeros((len(self.timestamp)), dtype=np.float32)
        self.length = len(self.data)
        
        # 填充数据到512Hz
        self.fill_data(target_sampling_rate=512)
        
        # 清理临时变量
        gc.collect()
        
        self.dataLvbo()
        # 优化：使用numpy数组，减少内存使用
        self.labels = np.zeros(self.length, dtype=int)
        self.calculateIndexlen()
        
        # 清理临时变量
        gc.collect()
        
        print(rf"完成读取 {self.filename} {datetime.now()}")



    def saveData(self):
        filename_without_ext = os.path.splitext(os.path.basename(self.filepath))[0]
        # 构建新的 HDF5 文件名
        h5py_filename = filename_without_ext + ".hdf"
        clean_h5py_filename = filename_without_ext+"_clean" + ".hdf"
        savepath=os.path.join(self.savepath,h5py_filename)

        labelsavedir=os.path.join(self.savepath,"labels")
        if not os.path.exists(labelsavedir):
            os.makedirs(labelsavedir)
        labelsavepath = os.path.join(labelsavedir, f"{self.filename}.csv")


        # 创建 HDF5 文件
        with h5py.File(savepath, "w") as f:
            # 创建数据集并写入数据
            f.create_dataset('timestamp', data=self.timestamp)
            f.create_dataset("data", data=self.data)
            f.create_dataset("labels", data=self.labels)
        if(self.cleanfile):
            cleandir = os.path.join(self.savepath, "cleanfile")
            clean_savepath = os.path.join(cleandir, clean_h5py_filename)
            if (not os.path.exists(cleandir)):
                os.makedirs(cleandir)
            with h5py.File(clean_savepath, "w") as f:
                # 创建数据集并写入数据
                f.create_dataset('timestamp', data=self.timestamp)
                f.create_dataset("data", data=self.cleandata)
                f.create_dataset("labels", data=self.labels)
        # 方法2：直接创建DataFrame后过滤（更高效简洁）
        df = pd.DataFrame({
            'timestamp': self.timestamp,
            'label': self.labels
        })
        df = df[df['label'] == 1]
        df.to_csv(labelsavepath, index=False)
        print(f"数据已保存到 {h5py_filename}")
        self.data=None
        self.alldata=None
        self.labels=None
        self.timestamp=None
    def getState(self):
        return f"{self.index}/{self.indexlen}\n"
    def __iter__(self):
        return self
    def _double_diff(self, signal):
        """双差分运算增强QRS波群特征（网页1核心算法）"""
        diff1 = np.diff(signal, n=1)
        diff2 = np.diff(diff1, n=1)
        return np.pad(diff2, (1, 1), 'edge')  # 保持长度一致
    def _static_normalize(self, signal):
        """基于硬件量程的静态标准化（您的核心经验）"""
        # 将int16原始数据转为浮点型
        signal = signal.astype(np.float32)
        # 硬件量程标准化（32767对应±10mV）
        # signal = (signal / 32767.0) * 100.0
        signal = np.clip(signal, -20, 20)
        # 双差分处理（保持原有特征增强）
        return self._double_diff(signal)
    def getData(self):
        if self.index >= self.indexlen:
            return None
        if(self.index%100==0):
            print(f"{self.index}/{self.indexlen},{datetime.now()},{self.filename}")
        # print(f"{self.index}/{self.indexlen},{datetime.now()},{self.filename}")
        index = self.index
        if(self.ye==1):
            if (index < self.indexlen - 3):
                start = self.index * 5120
                end = start+ self.neuronLength
            else:
                start = self.length - self.neuronLength
                end = self.length
        else:
            if(index<self.indexlen-1):
                start=self.index*self.neuronLength
                end=(index+1)*self.neuronLength
            else:
                start = self.length-self.neuronLength
                end=self.length
        self.index+=1
        xdata=self.data[start:end]
        xdata = self._static_normalize(xdata)
        xdata = xiaobo(xdata, sample_rate=512, duration=15360 // 512)
        xdata=lvbo(xdata,512)
        wavelet = ECGFeatureExtractor.extract_wavelet_features(xdata)

        indexlist=getjilabel(xdata,yuzhi1=4,yuzhi2=9)
        labels=self.labels[start:end]
        for i in range(5120):
            labels[self.neuronLength-5120+i]=0

        return {"data":xdata, "label":labels, "indexlist":indexlist,"start":start,"swt":wavelet}
        # return self.data[start:end],self.labels[start:end],index*5120

    def setlabel(self,startindex,labels,cleanecg):
        if not isinstance(labels, np.ndarray):
            labels = np.array(labels)
        window_len = self.neuronLength  # 30秒 (15360)
        overlap_len = 512 * 13  # 13秒 (6656) - 用历史标注替换的前段长度
        new_data_len = 5120  # 10秒 (5120) - 每次前进的新数据长度（写回区域）
        assert len(labels) == window_len
        peak_count = int(labels.sum())
        if peak_count < 5:
            print(f"Page:{startindex//5120} detect very low R-wave, count :{peak_count}", self.filename)
        rough_mask = labels.copy()
        if startindex > 0:
            # 用已确认的历史标注替换当前窗口的前段，保证连续性
            safe_history_len = min(overlap_len, len(self.labels) - startindex)
            if safe_history_len > 0:
                rough_mask[:safe_history_len] = self.labels[startindex: startindex + safe_history_len]

        end_len = min(startindex + window_len, len(self.cleandata))

        final_mask_window = rough_mask.copy()
        if self.correct:
            xdata = self.data[startindex:startindex + window_len]
            # 末尾窗口 ECG 可能不足 window_len，补零避免校正器 assert 失败
            if len(xdata) < window_len:
                xdata = np.pad(xdata, (0, window_len - len(xdata)))
            corrected_label, info = self.corrector.correct(xdata, rough_mask)
            if info["message"] == "success":
                final_mask_window = corrected_label

        self.cleandata[startindex:end_len] = cleanecg[:end_len - startindex]

        if startindex == 0:
            self.labels[0: end_len] = final_mask_window[:end_len]
        elif startindex + window_len >= len(self.labels):  # 最后一个窗口
            self.labels[startindex: end_len] = final_mask_window[:end_len - startindex]
        else:
            if self.ye == 1:
                # 只写回最后 new_data_len（10秒）的新数据区域
                update_start_global = end_len - new_data_len
                self.labels[update_start_global: end_len] = final_mask_window[window_len - new_data_len:window_len]
            else:
                self.labels[startindex: end_len] = final_mask_window[:end_len - startindex]
    def __len__(self):
        return self.indexlen
    def dataLvbo(self):
        if self.data is not None and len(self.data) > 0:
            # 1. 强制将 self.data 转换为 numpy 数组，彻底断开与 Pandas 的联系
            # 这是为了防止在 self.data = ... 赋值时触发 Pandas 的 Length Mismatch
            raw_input = np.array(self.data).astype(float)
            filtered_result = lvbo(raw_input, self.samplerate)

            # 3. 赋值回 self.data
            self.data = filtered_result
            
    def fill_data(self, target_sampling_rate=512):
        """填充数据到目标采样率 - 按日期分组处理"""
        if self.data is None or len(self.data) < 2:
            print(f"初始数据检查：data={self.data}, len={len(self.data) if self.data is not None else 0}")
            return

        print(f"开始填充数据到 {target_sampling_rate}Hz...")
        print(f"原始数据点数量: {len(self.data)}")
        print(f"原始时间戳数量: {len(self.timestamp)}")
        start_time = time.time()
        
        # 1. 解析时间戳并按日期分组（向量化，替代逐条 Python 解析）
        print("解析时间戳并按日期分组...")
        ts_series = pd.Series(self.timestamp)
        dt_index = pd.to_datetime(ts_series)                       # 批量解析
        unix_times = dt_index.astype(np.int64).values / 1e9        # 纳秒 → 秒 float64
        date_parts = ts_series.str[:10].values                     # "YYYY-MM-DD"
        data_arr = np.asarray(self.data, dtype=np.float32)

        date_groups = {}
        for date in np.unique(date_parts):
            mask = date_parts == date
            date_groups[date] = {
                'timestamps': unix_times[mask],
                'data': data_arr[mask],
            }

        print(f"按日期分组完成，共 {len(date_groups)} 天数据")
        for date in sorted(date_groups.keys()):
            print(f"  {date}: {len(date_groups[date]['timestamps'])} 个数据点")

        # 计算总数据点数量
        total_points = sum(len(group['timestamps']) for group in date_groups.values())

        # 2. 处理每个日期的数据
        all_target_times = []
        all_target_data = []
        
        ARTIFACT_THRESHOLD = 460  # 低于此值视为严重缺失，原样保留，后续开发处理

        # 辅助函数：处理单个秒段（上采样 or 降采样均走此路径）
        def process_second(sec, timestamps_sec, data_array, target_sampling_rate):
            mask = (timestamps_sec >= sec) & (timestamps_sec < sec + 1.0)
            sec_data = data_array[mask]
            sec_times = timestamps_sec[mask]

            target_times = np.linspace(sec,
                                       sec + (target_sampling_rate - 1) / target_sampling_rate,
                                       target_sampling_rate)

            if not self.origin_resample:
                # ── 均匀重采样：线性插值到均匀网格（符合 PhysioNet/AHA 标准）──
                # 点数不足或超过都走 np.interp，自动完成上采样/降采样
                if len(sec_data) < 2:
                    return target_times, np.zeros(target_sampling_rate, dtype=np.float32)
                order = np.argsort(sec_times)
                interp_data = np.interp(target_times, sec_times[order], sec_data[order]).astype(np.float32)
                return target_times, interp_data

            # ── 保留原始时间戳模式 ──
            if len(sec_data) < 1:
                return target_times, np.zeros(target_sampling_rate, dtype=np.float32)

            sorted_indices = np.argsort(sec_times)
            sec_times = sec_times[sorted_indices]
            sec_data = sec_data[sorted_indices]

            existing_points = len(sec_times)
            points_to_insert = target_sampling_rate - existing_points

            if points_to_insert <= 0:
                # 超采样：用线性插值降采样到均匀 target_sampling_rate 网格
                order = np.argsort(sec_times)
                interp_data = np.interp(target_times, sec_times[order], sec_data[order]).astype(np.float32)
                return target_times, interp_data

            start_time = sec
            end_time = sec + 0.999
            if abs(sec_times[0] - start_time) >= 0.001:
                diff = sec_data[1] - sec_data[0] if len(sec_data) >= 2 else 0
                sec_times = np.insert(sec_times, 0, start_time)
                sec_data = np.insert(sec_data, 0, sec_data[0] - diff)
            if abs(sec_times[-1] - end_time) >= 0.001:
                diff = sec_data[-1] - sec_data[-2] if len(sec_data) >= 2 else 0
                sec_times = np.append(sec_times, end_time)
                sec_data = np.append(sec_data, sec_data[-1] + diff)

            existing_points = len(sec_times)
            points_to_insert = target_sampling_rate - existing_points

            if points_to_insert <= 0:
                # 加完边界点后仍超过目标点数，降采样到均匀网格
                order = np.argsort(sec_times)
                interp_data = np.interp(target_times, sec_times[order], sec_data[order]).astype(np.float32)
                return target_times, interp_data

            intervals = existing_points - 1
            points_per_interval = points_to_insert // intervals
            extra_points = points_to_insert % intervals

            combined_times = list(sec_times)
            combined_data = list(sec_data)
            insert_positions, insert_times, insert_data = [], [], []

            for i in range(intervals):
                insert_count = points_per_interval + (1 if i < extra_points else 0)
                if insert_count > 0:
                    t0, t1 = sec_times[i], sec_times[i + 1]
                    d0, d1 = sec_data[i], sec_data[i + 1]
                    for j in range(1, insert_count + 1):
                        insert_positions.append(i + j)
                        insert_times.append(round(t0 + (t1 - t0) * j / (insert_count + 1), 3))
                        insert_data.append(d0 + (d1 - d0) * j / (insert_count + 1))

            for pos, tv, dv in reversed(list(zip(insert_positions, insert_times, insert_data))):
                combined_times.insert(pos, tv)
                combined_data.insert(pos, dv)

            total_points_needed = target_sampling_rate
            if len(combined_times) > total_points_needed:
                combined_times = combined_times[:total_points_needed]
                combined_data = combined_data[:total_points_needed]
            elif len(combined_times) < total_points_needed:
                last_t, last_d = combined_times[-1], combined_data[-1]
                while len(combined_times) < total_points_needed:
                    last_t = round(last_t + 0.001, 3)
                    combined_times.append(last_t)
                    combined_data.append(last_d)

            return np.array(combined_times), np.array(combined_data, dtype=np.float32)

        # 处理每个日期
        for date in sorted(date_groups.keys()):
            print(f"\n处理日期: {date}")
            
            # 获取该日期的数据
            timestamps_sec = np.array(date_groups[date]['timestamps'], dtype=np.float64)
            data_array = np.array(date_groups[date]['data'], dtype=np.float32)
            
            print(f"  该日期数据点数量: {len(timestamps_sec)}")
            
            # 按秒分组，统计每秒的数据点数
            seconds = timestamps_sec.astype(np.int64)
            unique_seconds, counts = np.unique(seconds, return_counts=True)
            
            # 三类秒段分类：
            #   artifact  < 460          : 严重缺失，原样保留（后续开发处理）
            #   exact     == target_rate : 点数恰好，原样保留
            #   process   其余（含 >512）: 上采样或降采样到 target_sampling_rate
            artifact_seconds = unique_seconds[counts < ARTIFACT_THRESHOLD]
            exact_seconds    = unique_seconds[counts == target_sampling_rate]
            fill_seconds     = unique_seconds[(counts >= ARTIFACT_THRESHOLD) & (counts != target_sampling_rate)]

            print(f"  伪迹秒段(<{ARTIFACT_THRESHOLD}点): {len(artifact_seconds)}")
            print(f"  需要处理的秒段(含超采样): {len(fill_seconds)}")
            print(f"  点数精确等于{target_sampling_rate}的秒段: {len(exact_seconds)}")

            # 点数恰好或严重缺失的秒段原样保留
            passthrough_mask = np.isin(seconds, np.concatenate([exact_seconds, artifact_seconds]))
            all_target_times.extend(timestamps_sec[passthrough_mask].tolist())
            all_target_data.extend(data_array[passthrough_mask].tolist())

            # 处理需要填充/降采样的秒段
            if len(fill_seconds) > 0:
                print(f"  处理需要填充的秒段...")
                fill_results = []
                
                # 使用多线程并行处理，根据系统核心数量自动调整线程数
                # 获取系统核心数量
                cpu_count = os.cpu_count()
                # 根据核心数量计算线程数
                if cpu_count < 10:
                    max_workers = cpu_count // 2
                else:
                    max_workers = int(cpu_count * 2 / 3)
                # 确保至少有1个线程
                max_workers = max(1, max_workers)
                print(f"  使用 {max_workers} 个线程进行并行处理")
                
                with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                    # 提交任务
                    future_to_sec = {executor.submit(process_second, sec, timestamps_sec, data_array, target_sampling_rate): sec for sec in fill_seconds}
                    
                    # 处理结果
                    for i, future in enumerate(concurrent.futures.as_completed(future_to_sec)):
                        sec = future_to_sec[future]
                        try:
                            result = future.result()
                            fill_results.append(result)
                        except Exception as e:
                            print(f"  处理秒段 {sec} 时出错: {e}")
                        
                        # 显示进度
                        if (i + 1) % max(1, len(fill_seconds) // 4) == 0:
                            progress = (i + 1) / len(fill_seconds) * 100
                            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                            print(f"  [{current_time}] 处理进度: {progress:.1f}%，已处理 {i + 1} 个秒段")
                
                # 处理填充结果
                for target_times, interpolated_data in fill_results:
                    all_target_times.extend(target_times.tolist())
                    all_target_data.extend(interpolated_data.tolist())
                
                # 清理临时变量
                del fill_results
                gc.collect()

            # 清理临时变量
            del timestamps_sec, data_array, fill_seconds, artifact_seconds, exact_seconds, passthrough_mask, seconds, unique_seconds, counts
            gc.collect()

        # 清理临时变量
        del date_groups
        gc.collect()

        # 3. 转换为numpy数组并排序
        print("\n合并结果...")
        filled_ts_float = np.array(all_target_times, dtype=np.float64)
        filled_data = np.array(all_target_data, dtype=np.float32)

        # 清理临时变量
        del all_target_times, all_target_data
        gc.collect()

        # 按时间戳排序（直接对float排序，无需解析字符串）
        print("  按时间戳排序...")
        sorted_indices = np.argsort(filled_ts_float, kind='stable')
        filled_ts_float = filled_ts_float[sorted_indices]
        filled_data = filled_data[sorted_indices]
        del sorted_indices
        gc.collect()

        # 批量格式化时间戳字符串（一次性转换）
        print("  格式化时间戳...")
        filled_timestamps = (
            pd.to_datetime(filled_ts_float, unit='s')
            .strftime('%Y-%m-%d %H:%M:%S.%f')
            .str[:-3]
            .values
        )
        del filled_ts_float
        gc.collect()

        # 4. 更新数据
        self.timestamp = filled_timestamps
        self.data = filled_data
        self.cleandata = np.zeros((len(self.timestamp)))
        self.length = len(self.data)
        self.calculateIndexlen()

        # 清理临时变量
        del filled_timestamps, filled_data
        gc.collect()

        end_time = time.time()
        print(f"填充完成，耗时: {end_time - start_time:.2f} 秒")
        print(f"填充完成，数据点从 {total_points} 增加到 {len(self.data)}")
        # 计算填充后的采样率
        if len(self.data) > 1:
            # 使用填充后的数据点数量除以估计的时间范围
            estimated_duration = len(self.data) / target_sampling_rate
            print(f"填充后采样率: {len(self.data) / estimated_duration:.2f} Hz")

def xiaobo(ecg_signal,sample_rate=512,duration=15):
    time = np.arange(0, duration, 1 / sample_rate)
    # 小波滤波
    wavelet_filtered_signal = wavelet_filter_ecg(ecg_signal, wavelet_name='db4', wavelet_level=4,
                                                 threshold_method='硬阈值')
    return wavelet_filtered_signal
def wavelet_filter_ecg(ecg_signal, wavelet_name='db4', wavelet_level=4, threshold_method='硬阈值', threshold_value=None):
    """
    对心电信号进行小波滤波降噪 (与之前代码相同)
    """
    wavelet = pywt.Wavelet(wavelet_name)
    coeffs = pywt.wavedec(ecg_signal, wavelet, level=wavelet_level)
    coeffs_thresholded = list(coeffs)
    for i in range(1, len(coeffs_thresholded)):
        detail_coeffs = coeffs_thresholded[i]
        if threshold_value is None:
            threshold = np.median(np.abs(detail_coeffs)) * 0.6745
        else:
            threshold = threshold_value
        if threshold_method == '硬阈值':
            coeffs_thresholded[i] = pywt.threshold(detail_coeffs, threshold, mode='hard')
        elif threshold_method == '软阈值':
            coeffs_thresholded[i] = pywt.threshold(detail_coeffs, threshold, mode='soft')
        else:
            raise ValueError("Threshold method must be '硬阈值' or '软阈值'")
    ecg_signal_filtered = pywt.waverec(coeffs_thresholded, wavelet)
    return ecg_signal_filtered
def butterBandPassFilter(lowcut, highcut, samplerate, order):
    "生成巴特沃斯带通滤波器"
    semiSampleRate = samplerate*0.5
    low = lowcut / semiSampleRate
    high = highcut / semiSampleRate
    b,a = signal.butter(order,[low,high],btype='bandpass')

    return b,a

def butterBandStopFilter(lowcut, highcut, samplerate, order):
    "生成巴特沃斯带阻滤波器"
    semiSampleRate = samplerate*0.5
    low = lowcut / semiSampleRate
    high = highcut / semiSampleRate
    b,a = signal.butter(order,[low,high],btype='bandstop')

    return b,a
def lvbo(x, iSampleRate):
    original_length = len(x)
    # 强制转为 numpy 数组，确保没有索引偏移
    x_arr = np.array(x).flatten()

    # 1. filtfilt 保证：零相位偏移（R波顶点不移动）
    b, a = butterBandPassFilter(3, 70, iSampleRate, order=4)
    x_arr = signal.filtfilt(b, a, x_arr)

    b, a = butterBandStopFilter(48, 52, iSampleRate, order=2)
    x_arr = signal.filtfilt(b, a, x_arr)

    # 2. 小波滤波处理
    # 小波变换通常是在信号末尾产生 padding 差异
    x_arr = wavelet_filter_ecg(x_arr, wavelet_name='db4', wavelet_level=4, threshold_method='硬阈值')

    # 3. 裁剪保证：只从末尾剪裁
    # 这样可以确保 Index 0 到 Index (original_length-1) 的相对位置绝对不动
    if len(x_arr) > original_length:
        # 只保留从 0 开始的原始长度部分
        x_arr = x_arr[:original_length]
    elif len(x_arr) < original_length:
        # 如果万一短了（极少见），在末尾补 0，不影响前面的位置
        x_arr = np.pad(x_arr, (0, original_length - len(x_arr)), 'constant')

    return x_arr
if __name__=="__main__":
    filepath="./Device2_BMD101Data_2025-09-23-21-57.txt"
    savepath="./"
    n=NeuronDataClass(filepath, savepath,cache=True)
    n.convertTocache()